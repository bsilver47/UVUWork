## This was written by Benjamin Silver and was last updated (aside from comments) in November 2023
## The purpose was to be run by a member of the ERP Software Services (ESS) team when preparing to implement updates
## The list of updates comes from Ellucian (expressed inline) and are tracked in an Excel sheet
## Instructions for the members of the team running this program are listed below the import list

## Playwright is used here to handle the navigation of Ellucian's website to download the .XLS file that Ellucian offers on their Updates screen
## A number of alternatives were considered in place of Playwright, Selenium being number one, but Playwright did the best at handling downloads at the time of writing this software
## Pandas is the most efficient program for writing to Excel sheets, generally, thus it is used for many of the Excel applications here
## The exception is where this program needs to append the new data to the existing Excel file, rather that rewrite it entirely
## This, specifically, is handled by Openpyxl, which is unable to process .XLS files as Pandas is
## Thus both Pandas and Openpyxl are used in order to handle different aspects of the Excel processing

from datetime import *
## will want to convert the next line to your username and password/file
from devcontainer import * #ellucian_username, ellucian_password ## This is used to import an unencrypted version of the encrypted username and password as saved in a different file
from math import ceil
import openpyxl
from playwright.sync_api import Page, expect
import asyncio
from playwright.async_api import async_playwright
import pandas as pd
from os import getcwd

## A crucial part of enabling this code to run is to:
## 1) make sure that the correct name of the excel file that is being written to is added into the master_path variable just below this block of comments and
## 2) make sure that the file opens to the sheet that should be written to
##      If 2) needs remedied, go to the page in question, save it, then close the file
## Also, if the Excel file that is being written to is saved to the cloud, please close the file before running this program as otherwise errors will result

current_location_path = getcwd()
print(current_location_path)
master_path = "Copy of Ellucian Releases (All Releases).xlsx"
ellucian_username = ellucian_username()
ellucian_password = ellucian_password()

def max_row_finder(sheet_object):
    verifier = True
    topicrow = 1
    while verifier:
        topicrow  += 1
        try:
            val_check = sheet_object.iloc[topicrow][1]
        except:
            verifier = False
            return topicrow

def max_column_finder(sheet_object):
    verifier = True
    topiccolumn = 1
    while verifier:
        topiccolumn += 1
        try:
            val_check = sheet_object.iloc[1][topiccolumn]
        except:
            topiccolumn -= 1
            verifier = False
            return topiccolumn

# This section organizes a list of products to exlude from this update
exclusion_sheet = pd.read_excel(master_path, sheet_name="Excluded Releases")
exclusion_max_row = max_row_finder(exclusion_sheet) - 1
exclusion_list = []
for localrow in range(exclusion_max_row + 1):
    item = exclusion_sheet.iloc[localrow][0]
    exclusion_list.append(item)

# Master is the name of the main sheet that is being written to
# This renders the sheet workable
master_object = openpyxl.load_workbook(master_path)
master_sheet_object = master_object.active
rrelname_col = 5
master_max_row = master_sheet_object.max_row
master_release_names = []
for row in range(master_max_row):
    localmasterrow = row + 2
    releaseval = master_sheet_object.cell(row=localmasterrow, column=rrelname_col).value
    if releaseval != None:
        master_release_names.append(releaseval)
    else:
        break
master_max_row = len(master_release_names)
master_max_column = 14

max_date = datetime(2022, 1, 1, 0, 0)
for day in range(2, master_max_row + 1):
    dtformat = '%m/%d/%Y'
    current_day = master_sheet_object.cell(row = day, column = 7).value
    if current_day == None:
        master_max_row = day - 1
        break
    if type(current_day) == str:
        current_day = datetime.strptime(current_day, dtformat)
    if max_date < current_day:
        max_date = current_day

# This function takes the given line from the downloaded file and adapts/writes it to the master file 
def row_writer(row):
    # "Product Line Product"
    master_sheet_object.cell(row=master_max_row, column=4).value = str(old_sheet_object.iloc[row]["Product Line"]) + " - " + str(old_sheet_object.iloc[row]["Product Name"])
    # "Release: Release Name"
    master_sheet_object.cell(row=master_max_row, column=5).value = str(old_sheet_object.iloc[row]["Release Name"])
    # "Target GA Date"
    master_sheet_object.cell(row=master_max_row, column=6).value = old_sheet_object.iloc[row]["Target GA Date"].strftime('%m/%d/%Y')
    # "Date Released"
    master_sheet_object.cell(row=master_max_row, column=7).value = old_sheet_object.iloc[row]["Date Released"].strftime('%m/%d/%Y')
    # "UVU Upgrade"
    master_sheet_object.cell(row=master_max_row, column=8).value = "future"
    # "Ellucian Status"
    master_sheet_object.cell(row=master_max_row, column=9).value = str(old_sheet_object.iloc[row]["State"])
    # "Summary"
    master_sheet_object.cell(row=master_max_row, column=10).value = str(old_sheet_object.iloc[row]["Summary"])
    # "Description"
    master_sheet_object.cell(row=master_max_row, column=11).value = str(old_sheet_object.iloc[row]["Description"])

# old is how this program refers to the file that is downloaded from the web portion
today = date.today()
olddownloadname = "EllucianUpdatesAsOf" + today.strftime('%m%d%y') + ".xls"
year_difference = today.year - max_date.year
month_difference = today.month - max_date.month

async def ellucianToExcel():
    async with async_playwright() as playwright:
        url = "http://login.ellucian.com"
        webkit = playwright.chromium
        browser = await webkit.launch(headless=False)
        context = await browser.new_context()
        page = await context.new_page()
        await context.clear_cookies()
        await page.goto(url)

        await page.get_by_label("username").fill(ellucian_username)
        await page.get_by_label("password").fill(ellucian_password)
        await page.get_by_role("button", name="Sign In").click()

        await page.wait_for_url("https://elluciansupport.service-now.com/customer_center")
        await page.get_by_label("Product Releases").click()
        
        await page.get_by_role("button", name="Show filter").click()

        await page.get_by_text("-- choose field --").click()
        await page.get_by_title("Date Released").click()
        await page.get_by_label("Filter operator").select_option("after")

        await page.get_by_label("-- None --. Date Chooser").click()

        if year_difference > 0:
           for y in range(year_difference * 12):
               await page.get_by_role("button", name="Previous month").click()
        if month_difference > 0:        
           for m in range(month_difference):
               await page.get_by_role("button", name="Previous month").click()   

        await page.get_by_role("gridcell", name=max_date.strftime('%B %d %Y %A')).click()
        await page.get_by_role("button", name="Run").dblclick()
        await page.wait_for_load_state()
        await page.wait_for_load_state()
        await page.wait_for_load_state()
        await page.wait_for_load_state()
        await page.get_by_label("Product Release Context Menu").click()
        async with page.expect_download() as download_info:
            await page.get_by_text("Export as Excel").click()
        download = await download_info.value

        await download.save_as(current_location_path + "/" + olddownloadname)

        await browser.close()    

resettool = 0
while resettool < 4:
    asyncio.run(ellucianToExcel())

    old_sheet_object = pd.read_excel(olddownloadname, sheet_name="Page 1")
    old_max_column = max_column_finder(old_sheet_object)
    old_max_row = max_row_finder(old_sheet_object)

    # this is the portion of the program that goes row by row to write from old file to the master file dependent on whether it is excluded or if it was already in the master file
    try:
        print("The update should have succeeded as of...")
        for row in range(old_max_row):
            error_not_found = True
            excluder = False
            running_name = old_sheet_object.iloc[row]["Release Name"]
            for i in exclusion_list:
                if running_name == i:
                    excluder = True
                    print(i)
                    break
            if excluder:
                pass
            elif running_name in master_release_names:
                row_locator = master_release_names.index(running_name) + 2
                master_sheet_object.delete_rows(row_locator)
                row_writer(row)
            else:
                master_max_row += 1
                row_writer(row)

        resettool = 7
        master_object.save(master_path)
        print("now.")
    except:
        print("There was an error. The program will run again.")
        resettool += 1
